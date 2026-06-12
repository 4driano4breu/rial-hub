# -*- coding: utf-8 -*-
"""Extração de dados de Excel e PDF — compartilhado por todos os módulos."""
import re
from pathlib import Path
from typing import Any

import openpyxl
import pdfplumber

from core.config import SECAO_PARA_CAT


def fmt(v: float) -> str:
    return "{:,.2f}".format(v).replace(",", "X").replace(".", ",").replace("X", ".")


def parse_br(s: str) -> float | None:
    s = s.strip().replace("\xa0", "").replace(" ", "")
    m = re.match(r"^(\d{1,3}(?:\.\d{3})*),(\d{1,3})$", s)
    if m:
        return float(s.replace(".", "").replace(",", "."))
    return None


def extrair_dados_xlsx(xlsx_path: str | Path) -> tuple[dict[str, str], dict[str, dict[str, float]]]:
    """
    Lê número da medição, período e breakdown por cidade de um .xlsx de medição.
    Retorna: (info, cidades)
      info    = {"num_medicao": str, "periodo": str}
      cidades = {NOME_UPPER: {"adm":f, "pav":f, "canteiro":f, "rocada":f, "terra":f, "total":f}}
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    ws_med = wb["Med"]
    rows_med = list(ws_med.iter_rows(values_only=True))

    # Número da medição (linha 7, col N)
    num_med: str = "?"
    try:
        val = rows_med[6][13]
        if isinstance(val, str) and ":" in val:
            num_med = val.split(":", 1)[1].strip()
        elif isinstance(val, str):
            num_med = val.strip()
    except (IndexError, TypeError):
        pass
    if num_med == "?":
        for row in rows_med[:10]:
            for cell in row:
                if isinstance(cell, str) and "Medi" in cell and ":" in cell:
                    num_med = cell.split(":", 1)[1].strip()
                    break
            if num_med != "?":
                break

    # Período (linha 3, col N)
    periodo: str = "?"
    try:
        val = rows_med[2][13]
        if isinstance(val, str):
            periodo = val.replace("Periodo:", "").strip()
    except (IndexError, TypeError):
        pass

    # Tabela de breakdown por cidade (aba Cidades)
    ws = wb["Cidades"]
    rows = list(ws.iter_rows(values_only=True))
    tbl_idx = None
    for i, row in enumerate(rows):
        if (
            len(row) > 36
            and isinstance(row[35], str) and "Municipio" in row[35]
            and isinstance(row[36], str) and "ADMINISTRA" in row[36].upper()
        ):
            tbl_idx = i
            break

    if tbl_idx is None:
        raise ValueError("Tabela de serviços não encontrada na aba 'Cidades'.")

    cidades: dict[str, dict[str, float]] = {}
    for row in rows[tbl_idx + 1: tbl_idx + 6]:
        if not row or len(row) < 42 or not row[35]:
            continue
        nome = row[35].strip().upper()
        d: dict[str, float] = {
            "adm":      float(row[36] or 0),
            "pav":      float(row[37] or 0),
            "canteiro": float(row[38] or 0),
            "rocada":   float(row[39] or 0),
            "terra":    float(row[40] or 0),
            "total":    float(row[41] or 0),
        }
        if d["total"] > 0:
            cidades[nome] = d

    if not cidades:
        raise ValueError("Nenhuma cidade com valor > 0 encontrada.")

    return {"num_medicao": num_med, "periodo": periodo}, cidades


def extrair_pdf_reajuste(pdf_path: str | Path) -> tuple[dict[str, float], float, dict[str, float]]:
    """
    Extrai do PDF de reajuste AGESUL:
      coefs_cat     : {"terra": 0.537, ...}
      total_pdf     : float  (valor total declarado)
      reaj_por_secao: {"04": 14089.46, ...}
    """
    FGV_NOMES: dict[str, list[str]] = {
        "adm":      ["Administração Local"],
        "terra":    ["Terraplenagem"],
        "canteiro": ["Índice Nacional", "INCC"],
        "rocada":   ["Conservação Rodoviária"],
        "pav":      ["Pavimentação"],
    }

    coefs_raw: dict[str, float] = {}
    total_pdf: float | None = None
    reaj_itens: dict[str, float] = {}

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""

            pat_fgv = (
                r"FGV\s+([\w\s\-ãõçáéíóúâêîôûÃÕÇÁÉÍÓÚ]+?)\s+"
                r"\w+\/\d{4}\s+\w+\/\d{4}\s+[\d,.]+\s+[\d,.]+\s+([\d,]+)%"
            )
            for m in re.finditer(pat_fgv, text):
                coefs_raw[m.group(1).strip()] = float(m.group(2).replace(",", ".")) / 100

            m_tot = re.search(
                r"VALOR DO REAJUSTE NO PER[IÍ]ODO[:\s]+([0-9.]+,[0-9]+)", text
            )
            if m_tot and total_pdf is None:
                total_pdf = parse_br(m_tot.group(1))

            pat_item = (
                r"\b(0[1-5])\.\d{3}(?:\.\d{3})?"
                r"[^\n\r]*?"
                r"([\d]{1,3}(?:\.[\d]{3})*,[\d]{2})"
                r"\s*\n"
            )
            for m in re.finditer(pat_item, text + "\n"):
                secao = m.group(1)
                val = parse_br(m.group(2))
                if val is not None and val > 0:
                    reaj_itens[secao] = reaj_itens.get(secao, 0.0) + val

    if not coefs_raw:
        raise ValueError("Tabela FGV não encontrada no PDF.")
    if total_pdf is None:
        raise ValueError("'VALOR DO REAJUSTE NO PERÍODO' não encontrado no PDF.")

    coefs_cat: dict[str, float] = {}
    for cat, nomes in FGV_NOMES.items():
        for k, v in coefs_raw.items():
            if any(n.lower() in k.lower() for n in nomes):
                coefs_cat[cat] = v
                break
        if cat not in coefs_cat:
            raise ValueError(
                f"Coeficiente para '{cat}' não encontrado. Disponíveis: {list(coefs_raw.keys())}"
            )

    return coefs_cat, total_pdf, reaj_itens

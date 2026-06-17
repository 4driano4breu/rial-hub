import os
import tempfile
from pathlib import Path
from flask import render_template, request, send_from_directory, current_app, flash, redirect, url_for
from flask_login import current_user

from app.blueprints.faturamento import faturamento_bp
from app.extensions import db

_R2_XLSX = "faturamento/Faturamento_2026.xlsx"
_R2_DASH = "faturamento/dashboard.html"


def _xlsx_path() -> Path:
    d = Path(current_app.instance_path) / "faturamento"
    d.mkdir(parents=True, exist_ok=True)
    return d / "Faturamento 2026.xlsx"


@faturamento_bp.route("/")
def index():
    from core.timestamps import ler_timestamps
    ts = ler_timestamps()
    return render_template("faturamento/index.html", ultima_atualizacao=ts.get("faturamento", "—"))


@faturamento_bp.route("/dashboard")
def dashboard():
    folder = Path(current_app.static_folder) / "ferramentas" / "faturamento"
    return send_from_directory(folder, "dashboard.html")


def regenerar_dashboard(notas_novas):
    import json
    import re
    from collections import defaultdict
    import app.storage as r2

    dash_path = Path(current_app.static_folder) / "ferramentas" / "faturamento" / "dashboard.html"
    if not dash_path.exists():
        return

    MESES_PT = {1:"JANEIRO",2:"FEVEREIRO",3:"MARÇO",4:"ABRIL",5:"MAIO",6:"JUNHO",
                7:"JULHO",8:"AGOSTO",9:"SETEMBRO",10:"OUTUBRO",11:"NOVEMBRO",12:"DEZEMBRO"}

    html = dash_path.read_text(encoding="utf-8")

    m = re.search(r'const NOTES = (\[.*?\]);', html, re.DOTALL)
    existing = json.loads(m.group(1)) if m else []
    existing_nrs = {n["nr"] for n in existing}

    for n in notas_novas:
        if n["nr"] not in existing_nrs:
            existing.append({
                "mes": MESES_PT[n["emissao"].month],
                "nr": n["nr"],
                "emissao": n["emissao"].strftime("%d/%m/%Y"),
                "contrato": n["contrato"],
                "orgao": n["orgao"],
                "municipio": n["municipio"],
                "tipo": n["tipo"],
                "bruto": round(n["bruto"], 2),
                "inss": round(n["inss"], 2),
                "ir": round(n["ir"], 2),
                "iss": round(n["iss"], 2),
                "liquido": round(n["liquido"], 2),
                "recebido": False,
                "dataRecebimento": "",
            })

    def sort_key(n):
        d = n["emissao"].split("/")
        return (int(d[2]), int(d[1]), -n["nr"])
    existing.sort(key=sort_key)

    month_data = defaultdict(lambda: {"notas": 0, "bruto": 0, "inss": 0, "ir": 0, "iss": 0, "liquido": 0})
    for n in existing:
        k = n["mes"]
        month_data[k]["notas"] += 1
        for f in ("bruto", "inss", "ir", "iss", "liquido"):
            month_data[k][f] += n[f]

    MESES_ORDER = list(MESES_PT.values())
    summary = []
    for mes in MESES_ORDER:
        if mes in month_data:
            d = month_data[mes]
            summary.append({
                "notas": d["notas"],
                "bruto": round(d["bruto"], 2),
                "inss": round(d["inss"], 2),
                "ir": round(d["ir"], 2),
                "iss": round(d["iss"], 2),
                "liquido": round(d["liquido"], 2),
                "mes": mes,
            })

    notes_js = json.dumps(existing, ensure_ascii=False, separators=(",", ":"))
    summary_js = json.dumps(summary, ensure_ascii=False, separators=(",", ":"))

    html = re.sub(r'const NOTES = \[.*?\];', f'const NOTES = {notes_js};', html, flags=re.DOTALL)
    html = re.sub(r'const SUMMARY = \[.*?\];', f'const SUMMARY = {summary_js};', html, flags=re.DOTALL)

    dash_path.write_text(html, encoding="utf-8")
    r2.upload(_R2_DASH, dash_path.read_bytes(), "text/html; charset=utf-8")


@faturamento_bp.route("/atualizar", methods=["POST"])
def atualizar():
    import app.storage as r2

    xml_file = request.files.get("xml")
    xlsx_file = request.files.get("xlsx")
    if not xml_file or not xml_file.filename.endswith(".xml"):
        flash("Envie um arquivo .xml de NFS-e.", "error")
        return redirect(url_for("faturamento.index"))

    xml_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xml")
    xml_tmp.write(xml_file.read())
    xml_tmp.close()
    xml_path = Path(xml_tmp.name)

    xlsx_path = _xlsx_path()

    if xlsx_file and xlsx_file.filename.endswith(".xlsx"):
        xlsx_bytes = xlsx_file.read()
        xlsx_path.write_bytes(xlsx_bytes)
        r2.upload(_R2_XLSX, xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    elif not xlsx_path.exists():
        # Tenta recuperar do R2 antes de falhar
        xlsx_bytes = r2.download(_R2_XLSX)
        if xlsx_bytes:
            xlsx_path.write_bytes(xlsx_bytes)
        else:
            flash("Envie também a planilha .xlsx de faturamento (primeira vez).", "error")
            xml_path.unlink(missing_ok=True)
            return redirect(url_for("faturamento.index"))

    try:
        from app.blueprints.faturamento.updater import parse_xml, criar_aba_mes, atualizar_resumo
        from openpyxl import load_workbook

        notas = parse_xml(str(xml_path))
        if not notas:
            flash("Nenhuma nota encontrada no XML.", "error")
            return redirect(url_for("faturamento.index"))

        mes = notas[0]["emissao"].month
        ano = notas[0]["emissao"].year
        MESES_PT = {1:"JANEIRO",2:"FEVEREIRO",3:"MARÇO",4:"ABRIL",5:"MAIO",6:"JUNHO",
                    7:"JULHO",8:"AGOSTO",9:"SETEMBRO",10:"OUTUBRO",11:"NOVEMBRO",12:"DEZEMBRO"}
        nome_aba = f"{MESES_PT[mes]} {ano}"

        wb = load_workbook(str(xlsx_path))
        abas_mes = [s for s in wb.sheetnames if s != "RESUMO"]
        template_aba = abas_mes[-1]
        subtotal_row, col_map = criar_aba_mes(wb, notas, mes, ano, nome_aba, template_aba)
        atualizar_resumo(wb, mes, subtotal_row, col_map, nome_aba)
        wb.save(str(xlsx_path))

        # Persiste xlsx atualizado no R2
        r2.upload(_R2_XLSX, xlsx_path.read_bytes(),
                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Persiste notas no banco PostgreSQL
        from app.models import FaturamentoNota
        org_id = current_user.org_id if current_user.is_authenticated else 1
        for n in notas:
            exists = FaturamentoNota.query.filter_by(org_id=org_id, nr=n["nr"]).first()
            if not exists:
                emissao = n["emissao"].date() if hasattr(n["emissao"], "date") else n["emissao"]
                db.session.add(FaturamentoNota(
                    org_id=org_id,
                    nr=n["nr"],
                    emissao=emissao,
                    contrato=n.get("contrato", ""),
                    orgao=n.get("orgao", ""),
                    municipio=n.get("municipio", ""),
                    tipo=n.get("tipo", ""),
                    bruto=n.get("bruto", 0),
                    inss=n.get("inss", 0),
                    ir=n.get("ir", 0),
                    iss=n.get("iss", 0),
                    liquido=n.get("liquido", 0),
                ))
        db.session.commit()

        from core.timestamps import salvar_timestamp
        salvar_timestamp("faturamento")

        try:
            regenerar_dashboard(notas)
        except Exception as regen_err:
            flash(f"Planilha atualizada, mas erro ao regenerar dashboard: {regen_err}", "warning")
            return redirect(url_for("faturamento.index"))

        flash(f"Planilha e dashboard atualizados: {len(notas)} notas de {MESES_PT[mes]}/{ano} importadas.", "ok")
    except Exception as e:
        flash(f"Erro ao processar: {e}", "error")
    finally:
        xml_path.unlink(missing_ok=True)

    return redirect(url_for("faturamento.index"))

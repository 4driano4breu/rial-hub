#!/usr/bin/env python3
"""
Script para atualizar a planilha de Faturamento com dados do XML exportado da Prefeitura.
Uso: python atualizar_faturamento.py <arquivo.xml> [planilha.xlsx]
"""

import sys
import io
import re
import xml.etree.ElementTree as ET
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

NS = 'http://www.abrasf.org.br/nfse.xsd'

MESES_PT = {
    1: 'JANEIRO', 2: 'FEVEREIRO', 3: 'MARÇO',
    4: 'ABRIL', 5: 'MAIO', 6: 'JUNHO',
    7: 'JULHO', 8: 'AGOSTO', 9: 'SETEMBRO',
    10: 'OUTUBRO', 11: 'NOVEMBRO', 12: 'DEZEMBRO'
}

# -------------------------------------------------------------------
# Extração de campos da Discriminação
# -------------------------------------------------------------------

def extrair_contrato(discriminacao):
    padroes = [
        r'(?i)contrato\s+de\s+[\w\s]{0,40}n[º°\.]\s*([\d][\d./\-]+)',
        r'(?i)contrato\s+n[º°\.]\s*([\d][\d./\-]+)',
        r'(?i)contrato\s+[Nn][°º\.]\s+([\d][\d./\-]+)',
        r'(?i)contrato\s+([\d][\d./\-]+)',
        r'(?i)contrato\s+n[°º\.]?\s+([\d./\-]+)',
    ]
    for p in padroes:
        m = re.search(p, discriminacao)
        if m:
            val = m.group(1).strip().rstrip('.').rstrip('-')
            if val and re.search(r'\d', val):
                return val
    return ''


def extrair_municipio(discriminacao):
    m = re.search(
        r'MUN[IÍ]C[IÍ]PIO\s+DE\s+([A-ZÁÉÍÓÚÂÊÎÔÛÃÕÀÜÇ][A-ZÁÉÍÓÚÂÊÎÔÛÃÕÀÜÇ\s]+?)'
        r'(?:\s*[-/]\s*MS|\s*\.|\n|,|\s{2})',
        discriminacao, re.IGNORECASE
    )
    if m:
        return m.group(1).strip().upper()
    return 'CAMPO GRANDE'


def extrair_conta(discriminacao):
    bancos = ['SICREDI', 'BRADESCO', 'ITAU', 'SANTANDER',
              'BANCO DO BRASIL', 'CAIXA ECONOMICA', 'NUBANK']
    for banco in bancos:
        if banco.upper() in discriminacao.upper():
            return banco
    return ''


# -------------------------------------------------------------------
# Parse do XML
# -------------------------------------------------------------------

def _parse_valor_br(texto):
    """Converte '61.476,68' (formato BR) para float 61476.68."""
    if not texto or not texto.strip():
        return 0.0
    return float(texto.strip().replace('.', '').replace(',', '.'))


def parse_xml(xml_path):
    """Suporta formato listaNotaFiscalList exportado pela Prefeitura."""
    tree = ET.parse(xml_path)
    root = tree.getroot()
    notas = []

    items = root.findall('notaFiscalList')
    if not items:
        items = root.findall('.//notaFiscalList')

    for item in items:
        situacao = (item.findtext('situacaoDescription') or '').strip().lower()
        if situacao and situacao != 'ativa':
            continue

        nr_str = (item.findtext('numeroNota') or '').strip()
        nr = int(nr_str) if nr_str.isdigit() else 0
        if not nr:
            continue

        emissao_str = (item.findtext('dataHoraEmissao') or '').strip()
        emissao = None
        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y'):
            try:
                emissao = datetime.strptime(emissao_str, fmt)
                break
            except ValueError:
                pass

        orgao       = (item.findtext('nomeEmpresarial') or '').strip()
        valor_bruto = _parse_valor_br(item.findtext('valorServico'))
        iss         = _parse_valor_br(item.findtext('valorIssqnCalculado'))
        inss        = 0.0
        ir          = 0.0

        notas.append({
            'emissao':     emissao,
            'nr':          nr,
            'contrato':    '',
            'municipio':   '',
            'conta':       '',
            'orgao':       orgao,
            'tipo':        'PRINCIPAL',
            'valor_bruto': valor_bruto,
            'inss':        inss,
            'ir':          ir,
            'iss':         iss,
            'liquido':     valor_bruto - iss,
        })

    notas.sort(key=lambda x: x['emissao'] or datetime.min, reverse=True)
    return notas


# -------------------------------------------------------------------
# Carrega a planilha a partir de bytes (filesystem efêmero / R2)
# -------------------------------------------------------------------

def carregar_workbook(dados: bytes):
    return load_workbook(io.BytesIO(dados))


# -------------------------------------------------------------------
# Helpers de estilo
# -------------------------------------------------------------------

def thin_side():
    return Side(border_style='thin', color='000000')

def thin_border():
    s = thin_side()
    return Border(left=s, right=s, top=s, bottom=s)

def copy_cell_style(src, dst):
    if src.has_style:
        dst.font      = copy(src.font)
        dst.fill      = copy(src.fill)
        dst.border    = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


# -------------------------------------------------------------------
# Criação da aba
# -------------------------------------------------------------------

def criar_aba_mes(wb, notas, mes, ano, nome_aba, template_aba):
    # Remove aba existente com o mesmo nome
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]

    ws_tpl = wb[template_aba]

    # Lê estilos das linhas de cabeçalho e de uma linha de dados do template
    # para replicar na nova aba
    header_row_tpl = None
    data_row_tpl   = None
    subtotal_row_tpl = None

    for row in ws_tpl.iter_rows():
        for cell in row:
            v = str(cell.value or '').strip().upper()
            if 'EMISSÃO' in v or 'EMISSAO' in v:
                header_row_tpl = cell.row
            if 'SUBTOTAL' in v:
                subtotal_row_tpl = cell.row
    if header_row_tpl:
        data_row_tpl = header_row_tpl + 1

    # Descobre mapeamento de colunas no template
    col_map = {}
    if header_row_tpl:
        for cell in ws_tpl[header_row_tpl]:
            v = str(cell.value or '').strip().upper()
            if 'EMISSÃO' in v or 'EMISSAO' in v: col_map['emissao']    = cell.column
            elif 'Nº' in v or 'NOTA' in v:        col_map['nr']         = cell.column
            elif 'CONTRATO' in v:                  col_map['contrato']   = cell.column
            elif 'MUNI' in v or 'NICIPIO' in v or 'MNIC' in v: col_map['municipio'] = cell.column
            elif 'CONTA' in v:                     col_map['conta']      = cell.column
            elif 'ORGÃO' in v or 'ÓRGÃO' in v or 'ORGAO' in v: col_map['orgao'] = cell.column
            elif 'TIPO' in v:                      col_map['tipo']       = cell.column
            elif 'VALOR BRUTO' in v:               col_map['valor_bruto']= cell.column
            elif v == 'INSS':                      col_map['inss']       = cell.column
            elif v == 'IR':                        col_map['ir']         = cell.column
            elif v == 'ISS':                       col_map['iss']        = cell.column
            elif 'LÍQUIDO' in v or 'LIQUIDO' in v: col_map['liquido']   = cell.column
            elif 'SITUAÇÃO' in v or 'SITUACAO' in v: col_map['situacao']= cell.column
            elif 'PAG' in v:                       col_map['pagamento']  = cell.column

    # Cria nova aba vazia (SEM copiar o worksheet — evita arrastar tabelas)
    ws = wb.create_sheet(nome_aba)

    # Move para antes do RESUMO
    idx_resumo = wb.sheetnames.index('RESUMO')
    wb.move_sheet(ws, offset=idx_resumo - wb.sheetnames.index(nome_aba))

    # Copia larguras de colunas do template
    for col_letter, col_dim in ws_tpl.column_dimensions.items():
        ws.column_dimensions[col_letter].width = col_dim.width

    # Copia linhas de título e cabeçalho (acima do header de colunas + o próprio header)
    for r in range(1, header_row_tpl + 1):
        for cell in ws_tpl[r]:
            nc = ws.cell(row=r, column=cell.column)
            # Copia valor (exceto fórmulas de tabela estruturada)
            val = cell.value
            if isinstance(val, str) and val.startswith('=') and 'Tabela' in val:
                val = None
            nc.value = val
            copy_cell_style(cell, nc)

    # Copia linha de SUBTOTAL do template como referência de estilo
    if subtotal_row_tpl:
        for cell in ws_tpl[subtotal_row_tpl]:
            nc = ws.cell(row=subtotal_row_tpl, column=cell.column)
            copy_cell_style(cell, nc)

    # Define linhas de dados
    first_data_row = header_row_tpl + 1
    n = len(notas)
    subtotal_row_new = first_data_row + n

    # Captura estilo de uma linha de dados do template para replicar
    tpl_data_styles = {}
    if data_row_tpl and data_row_tpl < (subtotal_row_tpl or 999):
        for cell in ws_tpl[data_row_tpl]:
            tpl_data_styles[cell.column] = cell

    num_fmt_moeda = '#,##0.00'

    for i, nota in enumerate(notas):
        r = first_data_row + i

        def set_cell(campo, value, num_fmt=None):
            if campo not in col_map:
                return
            c = ws.cell(row=r, column=col_map[campo])
            c.value = value
            if num_fmt:
                c.number_format = num_fmt
            # Aplica estilo do template
            src = tpl_data_styles.get(col_map[campo])
            if src:
                copy_cell_style(src, c)
            c.border = thin_border()

        set_cell('emissao',     nota['emissao'],     'DD/MM/YYYY HH:MM:SS')
        set_cell('nr',          nota['nr'])
        set_cell('contrato',    nota['contrato'])
        set_cell('municipio',   nota['municipio'])
        set_cell('conta',       nota['conta'])
        set_cell('orgao',       nota['orgao'])
        set_cell('tipo',        nota['tipo'])
        set_cell('valor_bruto', nota['valor_bruto'], num_fmt_moeda)
        set_cell('inss',        nota['inss'],        num_fmt_moeda)
        set_cell('ir',          nota['ir'],          num_fmt_moeda)
        set_cell('iss',         nota['iss'],          num_fmt_moeda)

        # Valor Líquido — fórmula
        if 'liquido' in col_map:
            col_vb   = get_column_letter(col_map['valor_bruto'])
            col_inss = get_column_letter(col_map['inss'])
            col_ir   = get_column_letter(col_map['ir'])
            col_iss  = get_column_letter(col_map['iss'])
            c = ws.cell(row=r, column=col_map['liquido'])
            c.value = f'={col_vb}{r}-{col_inss}{r}-{col_ir}{r}-{col_iss}{r}'
            c.number_format = num_fmt_moeda
            src = tpl_data_styles.get(col_map['liquido'])
            if src:
                copy_cell_style(src, c)
            c.border = thin_border()

        # Colunas sem conteúdo também recebem borda
        for campo in ('situacao', 'pagamento'):
            if campo in col_map:
                c = ws.cell(row=r, column=col_map[campo])
                src = tpl_data_styles.get(col_map[campo])
                if src:
                    copy_cell_style(src, c)
                c.border = thin_border()

    # Linha de SUBTOTAL
    ws.cell(row=subtotal_row_new, column=col_map.get('emissao', 2)).value = 'SUBTOTAL'
    for campo in ('valor_bruto', 'inss', 'ir', 'iss', 'liquido'):
        if campo not in col_map:
            continue
        col_l = get_column_letter(col_map[campo])
        c = ws.cell(row=subtotal_row_new, column=col_map[campo])
        c.value = f'=SUBTOTAL(9,{col_l}{first_data_row}:{col_l}{first_data_row + n - 1})'
        c.number_format = num_fmt_moeda
        src = tpl_data_styles.get(col_map[campo])
        if src:
            copy_cell_style(src, c)
        c.border = thin_border()

    # Cria tabela do Excel para o range de dados (header + dados, sem subtotal)
    max_col = max(col_map.values())
    first_col_letter = get_column_letter(min(col_map.values()))
    last_col_letter  = get_column_letter(max_col)
    table_ref = f'{first_col_letter}{header_row_tpl}:{last_col_letter}{subtotal_row_new - 1}'
    table_name = f'Tabela{mes}{ano}'
    tbl = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)

    # Atualiza título
    for row in ws.iter_rows(max_row=header_row_tpl - 1):
        for cell in row:
            if cell.value and 'FATURAMENTO' in str(cell.value).upper():
                cell.value = f'FATURAMENTO {MESES_PT[mes]} {ano}'
                break

    print(f"  ✅ Aba '{nome_aba}' criada com {n} notas | Tabela: {table_ref}")
    return subtotal_row_new, col_map


# -------------------------------------------------------------------
# Atualiza RESUMO
# -------------------------------------------------------------------

def atualizar_resumo(wb, mes, subtotal_row, col_map, nome_aba):
    ws = wb['RESUMO']
    mes_nome = MESES_PT[mes]
    col_vb   = get_column_letter(col_map.get('valor_bruto', 9))
    col_inss = get_column_letter(col_map.get('inss', 10))
    col_iss  = get_column_letter(col_map.get('iss', 12))

    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value or '').strip().upper()
            if mes_nome in val:
                r = cell.row
                # Cabeçalho do bloco: procura na linha anterior
                for offset in [1, 2]:
                    header_r = r - offset
                    if header_r < 1:
                        continue
                    col_fat_r = col_inss_r = col_iss_r = None
                    for hcell in ws[header_r]:
                        hv = str(hcell.value or '').upper()
                        if 'FATURAMENTO' in hv: col_fat_r  = hcell.column
                        elif 'INSS' in hv:       col_inss_r = hcell.column
                        elif 'ISS' in hv:        col_iss_r  = hcell.column
                    if col_fat_r:
                        if col_fat_r:  ws.cell(row=r, column=col_fat_r ).value = f"='{nome_aba}'!{col_vb}{subtotal_row}"
                        if col_inss_r: ws.cell(row=r, column=col_inss_r).value = f"='{nome_aba}'!{col_inss}{subtotal_row}"
                        if col_iss_r:  ws.cell(row=r, column=col_iss_r ).value = f"='{nome_aba}'!{col_iss}{subtotal_row}"
                        print(f"  ✅ RESUMO atualizado para {mes_nome}")
                        return
    print(f"  ⚠️  '{mes_nome}' não encontrado no RESUMO.")


# -------------------------------------------------------------------
# Main
# -------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Uso: python atualizar_faturamento.py <arquivo.xml> [planilha.xlsx]")
        sys.exit(1)

    xml_path  = sys.argv[1]
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else 'Faturamento 2026.xlsx'

    print(f"\n📂 Lendo XML: {xml_path}")
    notas = parse_xml(xml_path)
    if not notas:
        print("Nenhuma nota encontrada no XML.")
        sys.exit(1)

    mes  = notas[0]['emissao'].month
    ano  = notas[0]['emissao'].year
    nome_aba = f"{MESES_PT[mes]} {ano}"
    print(f"   Encontradas {len(notas)} notas — {MESES_PT[mes]}/{ano}")

    print(f"\n📊 Abrindo planilha: {xlsx_path}")
    wb = load_workbook(xlsx_path)

    abas_mes = [s for s in wb.sheetnames if s != 'RESUMO']
    template_aba = abas_mes[-1]
    print(f"   Template: '{template_aba}'")

    print(f"\n✏️  Criando aba '{nome_aba}'...")
    subtotal_row, col_map = criar_aba_mes(wb, notas, mes, ano, nome_aba, template_aba)

    print(f"\n📋 Atualizando RESUMO...")
    atualizar_resumo(wb, mes, subtotal_row, col_map, nome_aba)

    out_path = xlsx_path.replace('.xlsx', '_atualizado.xlsx')
    if '_atualizado' in xlsx_path:
        out_path = xlsx_path
    wb.save(out_path)
    print(f"\n💾 Planilha salva: {out_path}")
    return out_path


if __name__ == '__main__':
    main()
